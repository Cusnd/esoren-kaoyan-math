from __future__ import annotations

import json
import hashlib
import os
import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path


ROOT = Path(__file__).resolve().parents[2]
SCRIPT = (
    ROOT
    / ".agents/skills/kaoyan-math1-fullscore-coach/scripts/extract_calculus_knowledge.py"
)
sys.path.insert(0, str(SCRIPT.parent))
import extract_calculus_knowledge as extractor  # noqa: E402


class CalculusKnowledgeExtractorTests(unittest.TestCase):
    def run_extractor(
        self, *identifiers: str, root: Path = ROOT
    ) -> subprocess.CompletedProcess[str]:
        environment = os.environ.copy()
        environment["PYTHONUTF8"] = "1"
        return subprocess.run(
            [
                sys.executable,
                str(SCRIPT),
                *identifiers,
                "--root",
                str(root),
                "--format",
                "json",
            ],
            text=True,
            encoding="utf-8",
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            env=environment,
            check=False,
        )

    def parse_success(self, *identifiers: str) -> dict[str, object]:
        completed = self.run_extractor(*identifiers)
        self.assertEqual(0, completed.returncode, completed.stderr)
        self.assertEqual("", completed.stderr)
        return json.loads(completed.stdout)

    def copy_resources(self, fixture: Path) -> Path:
        target = fixture / "resources"
        shutil.copytree(ROOT / "resources", target)
        return target

    def copy_registry_fixture(self, fixture: Path) -> None:
        self.copy_resources(fixture)
        target_data = fixture / "data"
        target_data.mkdir(parents=True)
        for name in ("knowledge_registry.yml", "textbook_catalog.yml"):
            shutil.copy2(ROOT / "data" / name, target_data / name)

    def test_k031_resolves_stable_node_anchor_and_t03(self) -> None:
        result = self.parse_success("K031")

        self.assertEqual(["K031"], result["ordered_ids"])
        item = result["items"][0]
        self.assertEqual("MATH1-KN-CALC-0080", item["stable_node"]["id"])
        self.assertEqual(["K031"], item["source_ids"])
        self.assertEqual("calc-01", item["chapter"]["chapter_key"])
        self.assertTrue(item["placement"]["existing_anchor"])
        self.assertTrue(item["placement"]["anchor_verified"])
        self.assertEqual(
            ["T03"],
            [family["source_id"] for family in item["related_problem_families"]],
        )

    def test_multiple_ids_are_extracted_in_one_result(self) -> None:
        result = self.parse_success("K031", "K032")

        self.assertEqual(["K031", "K032"], result["requested_ids"])
        self.assertEqual(
            {"K031", "K032"},
            {
                source_id
                for item in result["items"]
                for source_id in item["source_ids"]
            },
        )

    def test_shared_stable_node_keeps_k068_identity_and_is_emitted_once(self) -> None:
        single = self.parse_success("K068")
        self.assertEqual(["K068"], single["items"][0]["source_ids"])
        self.assertEqual("K068", single["items"][0]["stable_node"]["source_id"])

        combined = self.parse_success("K067", "K068")
        self.assertEqual(1, len(combined["items"]))
        self.assertEqual(["K067", "K068"], combined["items"][0]["source_ids"])
        self.assertEqual(
            ["K067", "K068"],
            [record["source_id"] for record in combined["items"][0]["source_records"]],
        )

    def test_unanchored_k046_reports_catalog_target_without_inventing_anchor(self) -> None:
        result = self.parse_success("K046")

        item = result["items"][0]
        self.assertFalse(item["placement"]["existing_anchor"])
        self.assertEqual(
            "tex/chapters/calculus/03_one_variable_differential_concepts.tex",
            item["placement"]["target_chapter_file"],
        )
        self.assertNotIn("anchor_id", item["placement"])

    def test_duplicate_identifier_fails_without_partial_json(self) -> None:
        completed = self.run_extractor("K031", "K031")

        self.assertEqual(2, completed.returncode)
        self.assertEqual("", completed.stdout)
        self.assertIn("Duplicate", completed.stderr)

    def test_invalid_identifier_fails_entire_mixed_request(self) -> None:
        completed = self.run_extractor("K031", "K999")

        self.assertEqual(2, completed.returncode)
        self.assertEqual("", completed.stdout)
        self.assertIn("K001-K262", completed.stderr)

    def test_personal_columns_are_omitted_but_research_planning_is_separate(self) -> None:
        completed = self.run_extractor("K031")
        self.assertEqual(0, completed.returncode, completed.stderr)
        serialized = completed.stdout
        for forbidden in ("学习状态", "个人笔记", "复盘次数", "最后复习日期", "个人备注"):
            self.assertNotIn(forbidden, serialized)
        result = json.loads(serialized)
        planning = result["items"][0]["source_records"][0]["research_planning"]
        self.assertEqual("S", planning["重要度"])
        self.assertEqual("高频", planning["历年频率"])
        self.assertTrue(
            result["publication_guardrails"]["research_planning_is_internal_only"]
        )

    def test_registered_hash_drift_fails_before_any_json_output(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            fixture = Path(temporary)
            target_resources = self.copy_resources(fixture)
            narrative = target_resources / "考研数学一高等数学全量知识点地图_2026.md"
            narrative.write_text(
                narrative.read_text(encoding="utf-8") + "\n漂移测试\n", encoding="utf-8"
            )

            completed = self.run_extractor("K031", root=fixture)

        self.assertEqual(1, completed.returncode)
        self.assertEqual("", completed.stdout)
        self.assertIn("hash mismatch", completed.stderr)

    def test_windows_drive_and_symlink_paths_cannot_escape_root(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            base = Path(temporary)
            root = base / "repo"
            outside = base / "outside"
            root.mkdir()
            outside.mkdir()
            (outside / "source.txt").write_text("secret", encoding="utf-8")

            with self.assertRaises(extractor.SourceIntegrityError):
                extractor._safe_registered_path(root, r"C:\Windows\win.ini")

            link = root / "linked"
            try:
                link.symlink_to(outside, target_is_directory=True)
            except OSError:
                return
            with self.assertRaises(extractor.SourceIntegrityError):
                extractor._safe_registered_path(root, "linked/source.txt")

    def test_anchor_must_be_one_uncommented_content_anchor(self) -> None:
        target_file = "tex/chapters/calculus/test.tex"
        node = {
            "id": "MATH1-KN-CALC-TEST",
            "tex_anchor": {"id": "MATH1-KN-CALC-TEST", "file": target_file},
        }
        invalid_contents = {
            "comment_only": "% \\knowledgeAnchor[MATH1-KN-CALC-TEST]{注释}\n",
            "index_only": "\\knowledgeIndexAnchor[MATH1-KN-CALC-TEST]{索引}\n",
            "duplicate": (
                "\\knowledgeAnchor[MATH1-KN-CALC-TEST]{一}\n"
                "\\knowledgeAnchor[MATH1-KN-CALC-TEST]{二}\n"
            ),
        }
        for label, content in invalid_contents.items():
            with self.subTest(label=label), tempfile.TemporaryDirectory() as temporary:
                root = Path(temporary)
                path = root / target_file
                path.parent.mkdir(parents=True)
                path.write_text(content, encoding="utf-8")
                with self.assertRaises(extractor.SourceIntegrityError):
                    extractor._anchor_status(root, node, target_file)

    def test_valid_k_with_missing_registry_mapping_is_source_integrity_failure(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            fixture = Path(temporary)
            self.copy_registry_fixture(fixture)
            registry = fixture / "data/knowledge_registry.yml"
            text = registry.read_text(encoding="utf-8")
            self.assertIn("calc-map-2026:K031", text)
            registry.write_text(
                text.replace("calc-map-2026:K031", "calc-map-2026:K999", 1),
                encoding="utf-8",
            )

            completed = self.run_extractor("K031", root=fixture)

        self.assertEqual(1, completed.returncode)
        self.assertEqual("", completed.stdout)
        self.assertIn("stable knowledge node", completed.stderr)

    def test_valid_k_with_duplicate_registry_mapping_is_source_integrity_failure(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            fixture = Path(temporary)
            self.copy_registry_fixture(fixture)
            registry = fixture / "data/knowledge_registry.yml"
            text = registry.read_text(encoding="utf-8")
            marker = "      - calc-map-2026:K032"
            self.assertIn(marker, text)
            registry.write_text(
                text.replace(marker, f"{marker}\n      - calc-map-2026:K031", 1),
                encoding="utf-8",
            )

            completed = self.run_extractor("K031", root=fixture)

        self.assertEqual(1, completed.returncode)
        self.assertEqual("", completed.stdout)
        self.assertIn("found 2", completed.stderr)

    def test_workbook_schema_error_is_sanitized_and_has_no_partial_json(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            fixture = Path(temporary)
            resources = self.copy_resources(fixture)
            workbook_path = next(resources.glob("*.xlsx"))
            old_hash = hashlib.sha256(workbook_path.read_bytes()).hexdigest()

            from openpyxl import load_workbook

            workbook = load_workbook(workbook_path)
            sheet = workbook["题型方法库"]
            for row in sheet.iter_rows(max_row=12):
                for cell in row:
                    if cell.value == "标准流程":
                        cell.value = None
                        break
                else:
                    continue
                break
            workbook.save(workbook_path)
            workbook.close()

            new_hash = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
            manifest = resources / "manifest.yml"
            manifest_text = manifest.read_text(encoding="utf-8")
            self.assertIn(old_hash, manifest_text)
            manifest.write_text(manifest_text.replace(old_hash, new_hash, 1), encoding="utf-8")

            completed = self.run_extractor("K031", root=fixture)

        self.assertEqual(1, completed.returncode)
        self.assertEqual("", completed.stdout)
        self.assertIn("missing required headers", completed.stderr)
        self.assertNotIn("Traceback", completed.stderr)
        self.assertNotIn(str(fixture), completed.stderr)


if __name__ == "__main__":
    unittest.main()
