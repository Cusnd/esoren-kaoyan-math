#!/usr/bin/env python3
"""Extract verified, source-scoped calculus evidence for K identifiers."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path, PurePosixPath, PureWindowsPath
from typing import Any, Iterable

from repo_model import (
    RepositoryDataError,
    RepositoryDependencyError,
    load_catalog,
    load_knowledge_registry,
    load_resource_manifest,
)


PACKAGE_ID = "calc-map-2026"
K_PATTERN = re.compile(r"^K(?P<number>\d{3})$", re.IGNORECASE)
SOURCE_REF_PATTERN = re.compile(rf"^{PACKAGE_ID}:(?P<id>[KT]\d{{2,3}})$")
EXPECTED_ROLES = {"narrative_map", "audit_checklist", "structured_workbook"}
EXPECTED_SOURCE_RANGES = {"knowledge": "K001-K262", "problem_families": "T01-T51"}
EXPECTED_EXCLUDED_PUBLIC_FIELDS = {
    "exam_frequency",
    "trend",
    "research_rating",
    "learning_status",
    "review_count",
    "review_date",
    "personal_notes",
}

KNOWLEDGE_TEACHING_FIELDS = (
    "教学章",
    "官方模块",
    "专题",
    "知识点",
    "子知识点/必会动作",
    "口径",
    "大纲要求",
    "定义/公式/结论",
    "适用条件",
    "常见直接题型",
    "综合方向",
    "识别信号",
    "易错点",
    "前置知识",
    "推荐掌握程度",
)
KNOWLEDGE_PLANNING_FIELDS = ("层级", "典型真题年份", "历年频率", "重要度", "趋势")
CHECKLIST_TEACHING_FIELDS = ("教学章", "知识点与必会动作", "口径")
CHECKLIST_PLANNING_FIELDS = ("重要度", "层级", "典型真题年份")
METHOD_TEACHING_FIELDS = (
    "题型名称",
    "典型特征",
    "识别信号",
    "标准流程",
    "可替代方法",
    "方法选择条件",
    "易失败方法",
    "常见变式",
    "组合章节",
)
METHOD_PLANNING_FIELDS = ("典型年份", "难度", "建议训练量")
SHEET_REQUIRED_HEADERS = {
    "知识点总表": set(KNOWLEDGE_TEACHING_FIELDS) | set(KNOWLEDGE_PLANNING_FIELDS) | {"编号"},
    "题型方法库": set(METHOD_TEACHING_FIELDS) | set(METHOD_PLANNING_FIELDS) | {"编号"},
    "查漏清单": set(CHECKLIST_TEACHING_FIELDS) | set(CHECKLIST_PLANNING_FIELDS) | {"编号"},
}
SHEET_REQUIRED_VALUES = {
    "知识点总表": {"编号", "教学章", "知识点", "口径", "定义/公式/结论"},
    "题型方法库": {"编号", "题型名称", "识别信号", "标准流程"},
    "查漏清单": {"编号", "教学章", "知识点与必会动作", "口径"},
}


class InputError(RuntimeError):
    """Raised for invalid or unsupported requested identifiers."""


class SourceIntegrityError(RuntimeError):
    """Raised when the registered research package has drifted."""


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _nonempty(value: Any) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def _selected_fields(row: dict[str, Any], names: Iterable[str]) -> dict[str, Any]:
    return {
        name: _json_value(row[name])
        for name in names
        if name in row and _nonempty(row[name])
    }


def _normalize_ids(values: list[str]) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for raw in values:
        value = raw.strip().upper()
        match = K_PATTERN.fullmatch(value)
        if match is None or not 1 <= int(match.group("number")) <= 262:
            raise InputError(f"Invalid calculus knowledge identifier: {raw!r}; expected K001-K262")
        if value in seen:
            raise InputError(f"Duplicate calculus knowledge identifier: {value}")
        seen.add(value)
        normalized.append(value)
    return normalized


def _research_package(manifest: dict[str, Any]) -> dict[str, Any]:
    packages = manifest.get("research_packages")
    if not isinstance(packages, list):
        raise SourceIntegrityError("resources/manifest.yml has no research_packages list")
    matches = [item for item in packages if isinstance(item, dict) and item.get("id") == PACKAGE_ID]
    if len(matches) != 1:
        raise SourceIntegrityError(f"{PACKAGE_ID} must appear exactly once in resources/manifest.yml")
    package = matches[0]
    if package.get("authority") != "non_authoritative_research" or package.get("build_input") is not False:
        raise SourceIntegrityError(f"{PACKAGE_ID} authority/build_input contract has changed")
    if package.get("source_ranges") != EXPECTED_SOURCE_RANGES:
        raise SourceIntegrityError(f"{PACKAGE_ID} source_ranges contract has changed")
    policy = package.get("projection_policy")
    if (
        not isinstance(policy, dict)
        or policy.get("searchable_aliases") is not True
        or policy.get("expose_source_refs") is not False
        or policy.get("import_personal_state") is not False
    ):
        raise SourceIntegrityError(f"{PACKAGE_ID} projection policy has changed")
    excluded = policy.get("excluded_public_fields")
    if not isinstance(excluded, list) or not EXPECTED_EXCLUDED_PUBLIC_FIELDS.issubset(excluded):
        raise SourceIntegrityError(f"{PACKAGE_ID} public-field exclusions are incomplete")
    return package


def _safe_registered_path(root: Path, value: Any) -> tuple[str, Path]:
    if not isinstance(value, str) or not value.strip():
        raise SourceIntegrityError("Registered path must be a non-empty string")
    normalized = value.replace("\\", "/")
    pure = PurePosixPath(normalized)
    windows = PureWindowsPath(value)
    if (
        pure.is_absolute()
        or windows.is_absolute()
        or bool(windows.drive)
        or ".." in pure.parts
        or any(":" in part for part in pure.parts)
    ):
        raise SourceIntegrityError(f"Unsafe registered path: {value!r}")
    root_resolved = root.resolve()
    resolved = (root_resolved / pure.as_posix()).resolve(strict=False)
    if not resolved.is_relative_to(root_resolved):
        raise SourceIntegrityError(f"Registered path escapes repository root: {value!r}")
    return pure.as_posix(), resolved


def _verified_files(root: Path, package: dict[str, Any]) -> dict[str, dict[str, Any]]:
    files = package.get("files")
    if not isinstance(files, list) or len(files) != len(EXPECTED_ROLES):
        raise SourceIntegrityError(f"{PACKAGE_ID} must register exactly three source files")
    verified: dict[str, dict[str, str]] = {}
    for item in files:
        if not isinstance(item, dict):
            raise SourceIntegrityError("Research package file entries must be mappings")
        role = item.get("role")
        expected_hash = item.get("sha256")
        if role not in EXPECTED_ROLES or role in verified:
            raise SourceIntegrityError(f"Unexpected or duplicate research package role: {role!r}")
        if not isinstance(expected_hash, str) or not re.fullmatch(r"[0-9a-f]{64}", expected_hash):
            raise SourceIntegrityError(f"Invalid SHA-256 for research package role {role}")
        relative, path = _safe_registered_path(root, item.get("path"))
        try:
            actual_hash = hashlib.sha256(path.read_bytes()).hexdigest()
        except OSError as exc:
            raise SourceIntegrityError(f"Cannot read registered source {relative}: {exc}") from exc
        if actual_hash != expected_hash:
            raise SourceIntegrityError(
                f"Registered source hash mismatch for {relative}: expected {expected_hash}, got {actual_hash}"
            )
        verified[role] = {"path": relative, "sha256": actual_hash, "resolved_path": path}
    if set(verified) != EXPECTED_ROLES:
        raise SourceIntegrityError(f"{PACKAGE_ID} source roles are incomplete")
    return verified


def _load_table(workbook: Any, sheet_name: str) -> dict[str, list[dict[str, Any]]]:
    if sheet_name not in workbook.sheetnames:
        raise SourceIntegrityError(f"Structured workbook is missing sheet {sheet_name!r}")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header_index: int | None = None
    headers: list[str | None] = []
    for index, row in enumerate(rows[:12]):
        candidate = [str(value).strip() if value is not None else None for value in row]
        if "编号" in candidate:
            header_index = index
            headers = candidate
            break
    if header_index is None:
        raise SourceIntegrityError(f"Sheet {sheet_name!r} has no 编号 header row")
    nonempty_headers = [header for header in headers if header is not None]
    duplicates = sorted({header for header in nonempty_headers if nonempty_headers.count(header) > 1})
    if duplicates:
        raise SourceIntegrityError(
            f"Sheet {sheet_name!r} has duplicate headers: {', '.join(duplicates)}"
        )
    missing_headers = sorted(SHEET_REQUIRED_HEADERS[sheet_name] - set(nonempty_headers))
    if missing_headers:
        raise SourceIntegrityError(
            f"Sheet {sheet_name!r} is missing required headers: {', '.join(missing_headers)}"
        )
    id_column = headers.index("编号")
    table: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for values in rows[header_index + 1 :]:
        if id_column >= len(values) or values[id_column] is None:
            continue
        identifier = str(values[id_column]).strip().upper()
        if not identifier:
            continue
        row = {
            header: _json_value(values[column])
            for column, header in enumerate(headers)
            if header is not None and column < len(values)
        }
        table[identifier].append(row)
    return dict(table)


def _one_row(table: dict[str, list[dict[str, Any]]], identifier: str, sheet: str) -> dict[str, Any]:
    matches = table.get(identifier, [])
    if len(matches) != 1:
        raise SourceIntegrityError(
            f"{sheet} must contain exactly one row for {identifier}; found {len(matches)}"
        )
    row = matches[0]
    missing_values = sorted(
        field for field in SHEET_REQUIRED_VALUES[sheet] if not _nonempty(row.get(field))
    )
    if missing_values:
        raise SourceIntegrityError(
            f"{sheet} row {identifier} has empty required fields: {', '.join(missing_values)}"
        )
    return row


def _load_workbook_tables(
    path: Path, display_path: str
) -> dict[str, dict[str, list[dict[str, Any]]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise RepositoryDependencyError(
            "openpyxl is required for calculus knowledge extraction; run in codex-tools"
        ) from exc
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        raise SourceIntegrityError(
            f"Cannot open structured workbook {display_path}: invalid workbook content ({type(exc).__name__})"
        ) from exc
    try:
        try:
            return {
                "知识点总表": _load_table(workbook, "知识点总表"),
                "题型方法库": _load_table(workbook, "题型方法库"),
                "查漏清单": _load_table(workbook, "查漏清单"),
            }
        except SourceIntegrityError:
            raise
        except Exception as exc:
            raise SourceIntegrityError(
                f"Cannot parse structured workbook {display_path}: invalid worksheet content ({type(exc).__name__})"
            ) from exc
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def _source_index(nodes: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    result: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for node in nodes:
        refs = node.get("source_refs", [])
        if not isinstance(refs, list):
            continue
        for ref in refs:
            match = SOURCE_REF_PATTERN.fullmatch(ref) if isinstance(ref, str) else None
            if match is not None:
                result[match.group("id")].append(node)
    return dict(result)


def _source_ids(node: dict[str, Any]) -> list[str]:
    refs = node.get("source_refs", [])
    if not isinstance(refs, list):
        return []
    result: list[str] = []
    for ref in refs:
        match = SOURCE_REF_PATTERN.fullmatch(ref) if isinstance(ref, str) else None
        if match is not None:
            result.append(match.group("id"))
    return result


def _source_id(node: dict[str, Any]) -> str | None:
    identifiers = _source_ids(node)
    return identifiers[0] if len(identifiers) == 1 else None


def _brief_node(
    node: dict[str, Any], matched_source_ids: list[str] | None = None
) -> dict[str, Any]:
    result = {
        "id": node.get("id"),
        "title": node.get("title"),
        "kind": node.get("kind"),
        "chapter_key": node.get("chapter_key"),
    }
    source_ids = matched_source_ids if matched_source_ids is not None else _source_ids(node)
    if len(source_ids) == 1:
        result["source_id"] = source_ids[0]
    elif source_ids:
        result["source_ids"] = source_ids
    return result


def _public_node(
    node: dict[str, Any], matched_source_ids: list[str] | None = None
) -> dict[str, Any]:
    result = _brief_node(node, matched_source_ids)
    aliases = node.get("aliases")
    if isinstance(aliases, list):
        result["aliases"] = aliases
    if node.get("reviewable") is not None:
        result["reviewable"] = node.get("reviewable")
    return result


def _strip_tex_comments(text: str) -> str:
    visible_lines: list[str] = []
    for line in text.splitlines():
        cut = len(line)
        for index, character in enumerate(line):
            if character != "%":
                continue
            backslashes = 0
            cursor = index - 1
            while cursor >= 0 and line[cursor] == "\\":
                backslashes += 1
                cursor -= 1
            if backslashes % 2 == 0:
                cut = index
                break
        visible_lines.append(line[:cut])
    return "\n".join(visible_lines)


def _anchor_status(root: Path, node: dict[str, Any], target_file: str) -> dict[str, Any]:
    anchor = node.get("tex_anchor")
    if anchor is None:
        return {
            "existing_anchor": False,
            "target_chapter_file": target_file,
        }
    if not isinstance(anchor, dict) or anchor.get("id") != node.get("id"):
        raise SourceIntegrityError(f"{node.get('id')}: malformed tex_anchor")
    relative, path = _safe_registered_path(root, anchor.get("file"))
    if PurePosixPath(relative).as_posix() != PurePosixPath(target_file).as_posix():
        raise SourceIntegrityError(
            f"{node.get('id')}: declared anchor file {relative} does not match catalog target {target_file}"
        )
    try:
        text = path.read_text(encoding="utf-8")
    except (OSError, UnicodeError) as exc:
        raise SourceIntegrityError(f"{node.get('id')}: cannot read declared anchor file {relative}: {exc}") from exc
    node_id = re.escape(str(node.get("id")))
    visible = _strip_tex_comments(text)
    matches = re.findall(rf"\\knowledgeAnchor\s*\[\s*{node_id}\s*\]", visible)
    if len(matches) != 1:
        raise SourceIntegrityError(
            f"{node.get('id')}: expected exactly one content anchor in {relative}; found {len(matches)}"
        )
    return {
        "existing_anchor": True,
        "content_file": relative,
        "anchor_id": node.get("id"),
        "anchor_verified": True,
        "target_chapter_file": target_file,
    }


def _matching_knowledge_excerpt(text: str, identifier: str) -> list[str]:
    preferred = [
        line.strip()
        for line in text.splitlines()
        if line.lstrip().startswith(f"- **{identifier}｜")
    ]
    if preferred:
        return preferred[:2]
    pattern = re.compile(rf"(?<![A-Z0-9]){re.escape(identifier)}(?!\d)")
    return [line.strip() for line in text.splitlines() if pattern.search(line)][:2]


def _matching_checklist_excerpt(text: str, identifier: str) -> list[str]:
    pattern = re.compile(rf"^- \[.\] {re.escape(identifier)}(?:\s|$)")
    return [line.strip() for line in text.splitlines() if pattern.search(line)][:2]


def _method_excerpt(text: str, identifier: str) -> list[str]:
    heading = re.compile(rf"^##\s+{re.escape(identifier)}(?:\s|$)")
    lines = text.splitlines()
    for index, line in enumerate(lines):
        if not heading.search(line):
            continue
        result: list[str] = []
        for candidate in lines[index : index + 40]:
            if result and candidate.startswith("## "):
                break
            if candidate.strip():
                result.append(candidate.strip())
        return result
    return []


def _ordered_requested_ids(
    requested: list[str], source_nodes: dict[str, list[dict[str, Any]]], edges: list[dict[str, Any]]
) -> tuple[list[str], str]:
    node_to_sources: dict[str, list[str]] = defaultdict(list)
    for source_id in requested:
        node_to_sources[source_nodes[source_id][0]["id"]].append(source_id)
    outgoing: dict[str, set[str]] = {source_id: set() for source_id in requested}
    indegree = {source_id: 0 for source_id in requested}
    for edge in edges:
        if not isinstance(edge, dict) or edge.get("type") != "prerequisite_for":
            continue
        sources = node_to_sources.get(edge.get("source"), [])
        targets = node_to_sources.get(edge.get("target"), [])
        for source in sources:
            for target in targets:
                if source == target or target in outgoing[source]:
                    continue
                outgoing[source].add(target)
                indegree[target] += 1
    original_index = {value: index for index, value in enumerate(requested)}
    ready = sorted((value for value in requested if indegree[value] == 0), key=original_index.get)
    ordered: list[str] = []
    while ready:
        value = ready.pop(0)
        ordered.append(value)
        for target in sorted(outgoing[value], key=original_index.get):
            indegree[target] -= 1
            if indegree[target] == 0:
                ready.append(target)
                ready.sort(key=original_index.get)
    if len(ordered) != len(requested):
        return requested, "user_order_due_to_relation_cycle"
    basis = "explicit_prerequisite_relations_then_user_order" if ordered != requested else "user_order"
    return ordered, basis


def extract(root: Path, requested: list[str]) -> dict[str, Any]:
    manifest = load_resource_manifest(root)
    package = _research_package(manifest)
    verified_files = _verified_files(root, package)
    workbook = verified_files["structured_workbook"]
    tables = _load_workbook_tables(workbook["resolved_path"], workbook["path"])

    registry = load_knowledge_registry(root)
    raw_nodes = registry.get("nodes")
    raw_edges = registry.get("edges", [])
    if not isinstance(raw_nodes, list) or not all(isinstance(node, dict) for node in raw_nodes):
        raise SourceIntegrityError("data/knowledge_registry.yml has no valid nodes list")
    if not isinstance(raw_edges, list) or not all(isinstance(edge, dict) for edge in raw_edges):
        raise SourceIntegrityError("data/knowledge_registry.yml has no valid edges list")
    nodes: list[dict[str, Any]] = raw_nodes
    edges: list[dict[str, Any]] = raw_edges
    nodes_by_id = {node.get("id"): node for node in nodes if isinstance(node.get("id"), str)}
    source_nodes = _source_index(nodes)

    for identifier in requested:
        matches = source_nodes.get(identifier, [])
        if len(matches) != 1:
            raise SourceIntegrityError(
                f"{identifier} must map to exactly one stable knowledge node; found {len(matches)}"
            )
        _one_row(tables["知识点总表"], identifier, "知识点总表")
        _one_row(tables["查漏清单"], identifier, "查漏清单")

    catalog = {entry.chapter_key: entry for entry in load_catalog(root)}
    ordered_ids, ordering_basis = _ordered_requested_ids(requested, source_nodes, edges)
    narrative_text = verified_files["narrative_map"]["resolved_path"].read_text(encoding="utf-8")
    checklist_text = verified_files["audit_checklist"]["resolved_path"].read_text(encoding="utf-8")

    grouped_ids: list[list[str]] = []
    grouped_node_ids: set[str] = set()
    for identifier in ordered_ids:
        node_id = source_nodes[identifier][0].get("id")
        if node_id in grouped_node_ids:
            continue
        grouped_node_ids.add(node_id)
        grouped_ids.append(
            [
                candidate
                for candidate in ordered_ids
                if source_nodes[candidate][0].get("id") == node_id
            ]
        )

    items: list[dict[str, Any]] = []
    for source_ids in grouped_ids:
        identifier = source_ids[0]
        node = source_nodes[identifier][0]
        chapter_key = node.get("chapter_key")
        chapter = catalog.get(chapter_key)
        if chapter is None or chapter.subject_key != "calculus":
            raise SourceIntegrityError(
                f"{', '.join(source_ids)}: unknown or non-calculus chapter {chapter_key!r}"
            )

        source_records: list[dict[str, Any]] = []
        for source_id in source_ids:
            knowledge_row = _one_row(tables["知识点总表"], source_id, "知识点总表")
            checklist_row = _one_row(tables["查漏清单"], source_id, "查漏清单")
            source_records.append(
                {
                    "source_id": source_id,
                    "teaching_evidence": _selected_fields(
                        knowledge_row, KNOWLEDGE_TEACHING_FIELDS
                    ),
                    "research_planning": _selected_fields(
                        knowledge_row, KNOWLEDGE_PLANNING_FIELDS
                    ),
                    "checklist_evidence": _selected_fields(
                        checklist_row, CHECKLIST_TEACHING_FIELDS
                    ),
                    "checklist_research_planning": _selected_fields(
                        checklist_row, CHECKLIST_PLANNING_FIELDS
                    ),
                    "research_excerpts_internal_only": {
                        "narrative_map": _matching_knowledge_excerpt(
                            narrative_text, source_id
                        ),
                        "audit_checklist": _matching_checklist_excerpt(
                            checklist_text, source_id
                        ),
                    },
                }
            )

        direct_relations: list[dict[str, Any]] = []
        related_family_ids: set[str] = set()
        for edge in edges:
            if edge.get("source") == node.get("id"):
                other = nodes_by_id.get(edge.get("target"))
                direction = "outgoing"
            elif edge.get("target") == node.get("id"):
                other = nodes_by_id.get(edge.get("source"))
                direction = "incoming"
            else:
                continue
            if other is None:
                raise SourceIntegrityError(
                    f"{', '.join(source_ids)}: relation references an unknown stable node"
                )
            direct_relations.append(
                {
                    "direction": direction,
                    "type": edge.get("type"),
                    "node": _brief_node(other),
                }
            )
            other_sources = _source_ids(other)
            if (
                direction == "outgoing"
                and edge.get("type") == "prerequisite_for"
                and other.get("kind") == "problem_family"
            ):
                related_family_ids.update(
                    source_id for source_id in other_sources if source_id.startswith("T")
                )

        related_families: list[dict[str, Any]] = []
        for family_id in sorted(related_family_ids):
            matches = source_nodes.get(family_id, [])
            if len(matches) != 1:
                raise SourceIntegrityError(
                    f"{family_id} must map to exactly one problem-family node; found {len(matches)}"
                )
            family_row = _one_row(tables["题型方法库"], family_id, "题型方法库")
            related_families.append(
                {
                    "source_id": family_id,
                    "stable_node": _public_node(matches[0], [family_id]),
                    "teaching_evidence": _selected_fields(family_row, METHOD_TEACHING_FIELDS),
                    "research_planning": _selected_fields(family_row, METHOD_PLANNING_FIELDS),
                    "research_excerpt_internal_only": _method_excerpt(narrative_text, family_id),
                }
            )

        direct_relations.sort(
            key=lambda item: (
                str(item["direction"]),
                str(item["type"]),
                str(
                    item["node"].get("source_id")
                    or ",".join(item["node"].get("source_ids", []))
                    or item["node"].get("id")
                ),
            )
        )
        items.append(
            {
                "source_ids": source_ids,
                "stable_node": _public_node(node, source_ids),
                "chapter": {
                    "subject_key": chapter.subject_key,
                    "subject_name": chapter.subject_name,
                    "number": chapter.number,
                    "chapter_key": chapter.chapter_key,
                    "title": chapter.title,
                    "file": chapter.file,
                },
                "placement": _anchor_status(root, node, chapter.file),
                "source_records": source_records,
                "direct_relations": direct_relations,
                "related_problem_families": related_families,
            }
        )

    return {
        "schema_version": 2,
        "package": {
            "id": package.get("id"),
            "title": package.get("title"),
            "authority": package.get("authority"),
            "build_input": package.get("build_input"),
            "verified_files": [
                {
                    "role": role,
                    "path": verified_files[role]["path"],
                    "sha256": verified_files[role]["sha256"],
                }
                for role in sorted(verified_files)
            ],
        },
        "publication_guardrails": {
            "teaching_evidence_may_be_paraphrased": True,
            "research_planning_is_internal_only": True,
            "source_claims_require_research_package_attribution": True,
            "personal_and_review_state_omitted": True,
            "source_refs_omitted": True,
        },
        "requested_ids": requested,
        "ordered_ids": ordered_ids,
        "ordering_basis": ordering_basis,
        "items": items,
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extract source-scoped calculus evidence for one or more K identifiers."
    )
    parser.add_argument("ids", nargs="+", help="One or more identifiers in K001-K262")
    parser.add_argument("--root", type=Path, default=Path.cwd())
    parser.add_argument("--format", choices=("json",), default="json")
    args = parser.parse_args()

    try:
        requested = _normalize_ids(args.ids)
        result = extract(args.root.resolve(), requested)
    except InputError as exc:
        print(str(exc), file=sys.stderr)
        return 2
    except RepositoryDependencyError as exc:
        print(str(exc), file=sys.stderr)
        return 2
    except (RepositoryDataError, SourceIntegrityError, OSError, UnicodeError) as exc:
        print(str(exc), file=sys.stderr)
        return 1

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
